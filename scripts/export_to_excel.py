# -*- coding: utf-8 -*-
"""
Script để chuyển đổi file CSV mô tả database thành file Excel (.xlsx)
với formatting đẹp và dễ đọc
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_from_csv(csv_file_path, output_excel_path):
    """
    Chuyển đổi file CSV thành Excel với formatting
    """
    try:
        # Đọc file CSV
        df = pd.read_csv(csv_file_path, encoding='utf-8')
        
        # Tạo workbook mới
        wb = Workbook()
        
        # Tạo worksheet chính
        ws_all = wb.active
        ws_all.title = "Tất cả bảng"
        
        # Định nghĩa style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Thêm dữ liệu vào worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_all.append(r)
        
        # Format header
        for cell in ws_all[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Format dữ liệu
        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-fit columns
        for column in ws_all.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_all.column_dimensions[column_letter].width = adjusted_width
        
        # Tạo các worksheet riêng cho từng bảng
        table_names = df['Tên Bảng'].unique()
        
        for table_name in table_names:
            # Tạo worksheet mới cho bảng
            ws_table = wb.create_sheet(title=table_name[:31])  # Excel giới hạn 31 ký tự cho tên sheet
            
            # Lọc dữ liệu cho bảng này
            table_data = df[df['Tên Bảng'] == table_name]
            
            # Thêm tiêu đề bảng
            ws_table.append([f"Mô tả chi tiết bảng: {table_name}"])
            ws_table.merge_cells('A1:I1')
            title_cell = ws_table['A1']
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Thêm dòng trống
            ws_table.append([])
            
            # Thêm header
            headers = list(table_data.columns)
            ws_table.append(headers)
            
            # Format header
            for cell in ws_table[3]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Thêm dữ liệu
            for _, row in table_data.iterrows():
                ws_table.append(list(row))
            
            # Format dữ liệu
            for row in ws_table.iter_rows(min_row=4, max_row=ws_table.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-fit columns cho worksheet này
            for column in ws_table.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_table.column_dimensions[column_letter].width = adjusted_width
        
        # Tạo worksheet tóm tắt
        ws_summary = wb.create_sheet(title="Tóm tắt", index=1)
        
        # Thêm thông tin tóm tắt
        summary_data = [
            ["THÔNG TIN TÓM TẮT CƠ SỞ DỮ LIỆU SUN MOVEMENT"],
            [],
            ["Tổng số bảng:", len(table_names)],
            ["Tổng số cột:", len(df)],
            [],
            ["DANH SÁCH CÁC BẢNG:"],
        ]
        
        for data_row in summary_data:
            ws_summary.append(data_row)
        
        # Thêm danh sách bảng với số cột
        for table_name in table_names:
            table_count = len(df[df['Tên Bảng'] == table_name])
            ws_summary.append([table_name, f"{table_count} cột"])
        
        # Format worksheet tóm tắt
        ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws_summary['A6'].font = Font(bold=True, size=12)
        
        # Format các dòng thông tin tóm tắt
        for row_num in [3, 4]:
            ws_summary[f'A{row_num}'].font = Font(bold=True)
        
        # Auto-fit columns
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Lưu file
        wb.save(output_excel_path)
        print(f"✅ Đã xuất thành công file Excel: {output_excel_path}")
        print(f"📊 Tổng số bảng: {len(table_names)}")
        print(f"📋 Tổng số cột: {len(df)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo file Excel: {str(e)}")
        return False

def main():
    """
    Hàm chính
    """
    # Đường dẫn file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    
    csv_file = os.path.join(project_root, "docs", "Database_Tables_Description.csv")
    excel_file = os.path.join(project_root, "docs", "Database_Tables_Description.xlsx")
    
    print("🚀 Bắt đầu chuyển đổi CSV sang Excel...")
    print(f"📁 File CSV nguồn: {csv_file}")
    print(f"📁 File Excel đích: {excel_file}")
    
    # Kiểm tra file CSV có tồn tại không
    if not os.path.exists(csv_file):
        print(f"❌ Không tìm thấy file CSV: {csv_file}")
        return
    
    # Chuyển đổi
    success = create_excel_from_csv(csv_file, excel_file)
    
    if success:
        print("\n✅ HOÀN THÀNH!")
        print(f"📄 File Excel đã được tạo tại: {excel_file}")
        print("\n📋 Cấu trúc file Excel:")
        print("  - Sheet 'Tóm tắt': Thông tin tổng quan")
        print("  - Sheet 'Tất cả bảng': Toàn bộ dữ liệu")
        print("  - Các sheet riêng: Mỗi bảng một sheet")
    else:
        print("\n❌ THẤT BẠI! Vui lòng kiểm tra lại.")

if __name__ == "__main__":
    main()
