import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Đường dẫn file
csv_file = r"d:\DATN\DATN\docs\Database_Tables_Description.csv"
excel_file = r"d:\DATN\DATN\docs\Database_Tables_Description.xlsx"

print("🚀 Bắt đầu chuyển đổi CSV sang Excel...")

try:
    # Đọc file CSV
    df = pd.read_csv(csv_file, encoding='utf-8')
    
    # Tạo workbook mới
    wb = Workbook()
    
    # Worksheet chính
    ws_all = wb.active
    ws_all.title = "Tất cả bảng"
    
    # Định nghĩa style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thêm dữ liệu vào worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_all.append(r)
    
    # Format header
    for cell in ws_all[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Format dữ liệu
    for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Set column widths manually
    column_widths = [5, 25, 25, 15, 10, 8, 8, 15, 50]
    for i, width in enumerate(column_widths, 1):
        ws_all.column_dimensions[chr(64 + i)].width = width
    
    # Tạo worksheet tóm tắt
    table_names = df['Tên Bảng'].unique()
    ws_summary = wb.create_sheet(title="Tóm tắt", index=0)
    
    # Thêm thông tin tóm tắt
    ws_summary.append(["THÔNG TIN TÓM TẮT CƠ SỞ DỮ LIỆU SUN MOVEMENT"])
    ws_summary.append([])
    ws_summary.append(["Tổng số bảng:", len(table_names)])
    ws_summary.append(["Tổng số cột:", len(df)])
    ws_summary.append([])
    ws_summary.append(["DANH SÁCH CÁC BẢNG:"])
    
    for table_name in table_names:
        table_count = len(df[df['Tên Bảng'] == table_name])
        ws_summary.append([table_name, f"{table_count} cột"])
    
    # Format worksheet tóm tắt
    ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_summary['A6'].font = Font(bold=True, size=12)
    
    # Set column widths cho summary
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    
    # Lưu file
    wb.save(excel_file)
    print(f"✅ Đã xuất thành công file Excel: {excel_file}")
    print(f"📊 Tổng số bảng: {len(table_names)}")
    print(f"📋 Tổng số cột: {len(df)}")
    
    # In danh sách bảng
    print("\n📋 Danh sách bảng trong database:")
    for i, table_name in enumerate(table_names, 1):
        table_count = len(df[df['Tên Bảng'] == table_name])
        print(f"  {i:2d}. {table_name} ({table_count} cột)")
    
except Exception as e:
    print(f"❌ Lỗi: {str(e)}")
    import traceback
    traceback.print_exc()
